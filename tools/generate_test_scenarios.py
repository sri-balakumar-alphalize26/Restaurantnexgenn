r"""Generate the master NexGenn POS Android APK test-scenario workbook.

Run:  python tools\generate_test_scenarios.py
Output:  docs\NexGenn_POS_Test_Scenarios.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "docs" / "NexGenn_POS_Test_Scenarios.xlsx"

# --- Columns ---------------------------------------------------------------
COLS = [
    ("TC ID",         8),
    ("Module",        18),
    ("Scenario",      45),
    ("Pre-condition", 35),
    ("Steps",         70),
    ("Expected Result", 55),
    ("Priority",      10),
    ("Type",          14),
    ("Automation",    28),
    ("Status",        14),
    ("Notes",         30),
]

# --- Scenarios -------------------------------------------------------------
SCENARIOS = [

    # ======================== 1. APP LAUNCH & DEVICE SETUP ========================
    ("TC-001", "App Launch", "Cold start opens Splash then Login",
        "APK freshly installed, no saved session",
        "1. Tap NexGenn POS icon\n2. Wait for splash\n3. Observe next screen",
        "Splash shows logo then Login screen with 'Welcome Back'",
        "P0", "Smoke", "Appium (home-and-orders.e2e.js covers post-login)", "Passed (manual 2026-04-24)", ""),
    ("TC-002", "App Launch", "Warm start after login shows Home",
        "User already logged in previously",
        "1. Kill app\n2. Reopen",
        "Skips Login, lands on Home dashboard",
        "P0", "Smoke", "Manual", "Passed (manual 2026-04-24)", "Session restored from AsyncStorage"),
    ("TC-003", "Device Setup", "First-run device registration",
        "Fresh install, no device_server_url",
        "1. Launch app\n2. Enter Odoo URL + DB\n3. Tap Save/Continue",
        "device_server_url + device_db_name persisted, Login screen shown",
        "P0", "Critical", "Manual", "Passed (manual 2026-04-24)", ""),
    ("TC-004", "Device Setup", "Device UUID generated and persisted",
        "Fresh install",
        "1. Complete device setup\n2. Reopen app after clearing cache but not data",
        "Same device_uuid stays in AsyncStorage across launches",
        "P1", "Regression", "Manual", "Not run", ""),
    ("TC-005", "Device Setup", "Gear icon removed from login header",
        "APK built after commit 17c4a53",
        "1. Open app to Login screen\n2. Inspect top-right header",
        "No gear/settings icon visible",
        "P0", "Regression", "Appium: smoke-login.e2e.js 'has no Device Setup gear icon'", "Passed (Appium 2026-04-24)", ""),

    # ======================== 2. LOGIN ==========================================
    ("TC-010", "Login", "Login with valid credentials",
        "Device config done, valid Odoo user",
        "1. Enter username\n2. Enter password\n3. Tap Login",
        "Redirects to Home dashboard; userData + session_id saved",
        "P0", "Smoke", "Appium (needs fresh Login state)", "Passed (manual admin/admin 2026-04-24)", ""),
    ("TC-011", "Login", "Login with invalid password",
        "Device config done",
        "1. Enter valid username\n2. Enter wrong password\n3. Tap Login",
        "Toast/Error: 'Invalid credentials'; stays on Login screen",
        "P0", "Critical", "Appium (needs fresh Login state)", "Not run", ""),
    ("TC-012", "Login", "Validate empty username and password",
        "-",
        "1. Leave username empty\n2. Leave password empty\n3. Tap Login",
        "Inline error shown for both fields (login-username-error / login-password-error)",
        "P1", "Critical", "Appium: smoke-login.e2e.js", "Blocked (Login precondition)", ""),
    ("TC-013", "Login", "Password visibility toggle",
        "-",
        "1. Type password\n2. Tap eye icon\n3. Tap again",
        "Password text toggles between masked and plain",
        "P2", "UI", "Appium: login-password-toggle", "Blocked (Login precondition)", ""),
    ("TC-014", "Login", "Autofill credentials toggle ON",
        "saved_credentials exist in AsyncStorage",
        "1. Toggle Autofill ON\n2. Observe fields",
        "Username + password auto-filled; toast 'Credentials filled'",
        "P1", "Functional", "Appium (needs fresh Login state)", "Not run", ""),
    ("TC-015", "Login", "Autofill toggle OFF clears fields",
        "Credentials filled via autofill",
        "1. Toggle Autofill OFF",
        "Both fields cleared",
        "P1", "Functional", "Manual", "Not run", ""),
    ("TC-016", "Login", "Server/KOT config primed after login",
        "valid user",
        "1. Login successfully\n2. Inspect AsyncStorage",
        "pos_config_id, kot_printer_ip, kot_printer_port cached",
        "P1", "Integration", "Manual", "Not run", ""),

    # ======================== 3. LANGUAGE / I18N =================================
    ("TC-020", "i18n", "Switch UI to Arabic (RTL)",
        "Login screen",
        "1. Tap AR (عربي) toggle\n2. Inspect layout",
        "Text translated; layout flips to RTL",
        "P1", "Critical", "Appium: smoke-login.e2e.js", "Not run", ""),
    ("TC-021", "i18n", "Switch back to English (LTR)",
        "Arabic active",
        "1. Tap EN toggle",
        "English strings restored; layout LTR",
        "P1", "Critical", "Manual", "Not run", ""),
    ("TC-022", "i18n", "Language choice persists across app restart",
        "Arabic selected",
        "1. Kill app\n2. Reopen",
        "App opens in Arabic",
        "P2", "Regression", "Manual", "Not run", ""),

    # ======================== 4. HOME / DASHBOARD ===============================
    ("TC-030", "Home", "Home shows Take Orders banner",
        "Logged in",
        "1. Land on Home",
        "'Take Orders' banner visible and tappable",
        "P0", "Smoke", "Appium: home-and-orders.e2e.js", "Passed (Appium 2026-04-24)", ""),
    ("TC-031", "Home", "Home shows Our Specials with Food + Drinks",
        "Logged in",
        "1. Scroll Home section",
        "Food and Drinks cards visible under 'Our Specials'",
        "P1", "Smoke", "Appium: home-and-orders.e2e.js", "Passed (Appium 2026-04-24)", ""),
    ("TC-032", "Home", "Bottom nav: Home / Profile / Logout present",
        "Logged in",
        "1. Inspect bottom bar",
        "All 3 tabs visible; Home highlighted",
        "P1", "UI", "Appium: home-and-orders.e2e.js", "Passed (Appium 2026-04-24)", ""),

    # ======================== 5. POS REGISTER / SESSION =========================
    ("TC-040", "POS Register", "Take Orders opens POS Register",
        "Logged in, active pos.session exists",
        "1. Tap 'Take Orders'",
        "POS Register screen lists Active Sessions with opening amounts",
        "P0", "Smoke", "Appium: home-and-orders 'opens Take Orders banner'", "Passed (Appium 2026-04-24)", ""),
    ("TC-041", "POS Register", "Continue Selling enters session",
        "POS Register shown",
        "1. Tap 'Continue Selling' on a session",
        "Product catalog screen opens; session context preserved",
        "P0", "Critical", "Manual", "Passed (manual 2026-04-24)", "User reached cart screen with Barotta + LED items"),
    ("TC-042", "POS Register", "Close Session",
        "POS Register shown",
        "1. Tap 'Close' on a session\n2. Confirm",
        "Session moves to closed state in Odoo",
        "P1", "Critical", "Manual", "Not run", ""),
    ("TC-043", "POS Register", "Register # matches Odoo config.name",
        "POS Register shown",
        "1. Compare Restaurant/00004 vs Odoo backend",
        "Register number identical",
        "P2", "Integration", "Manual", "Passed (visual 2026-04-24)", "Restaurant/00004 matched"),

    # ======================== 6. PRODUCT CATALOG ================================
    ("TC-050", "Catalog", "Food category shows products from Odoo",
        "Active session",
        "1. Tap Food card on Home",
        "Barotta, Fried Rice etc. listed with prices in OMR",
        "P0", "Smoke", "Appium: home-and-orders 'opens Food special card'", "Passed (Appium 2026-04-24)", "Barotta 1.000 / Fried Rice 4.000 / LED 110.000 OMR"),
    ("TC-051", "Catalog", "Product search by name",
        "Catalog open",
        "1. Type 'Bar' in search bar",
        "Results filtered to 'Barotta' only",
        "P1", "Functional", "Manual", "Not run", ""),
    ("TC-052", "Catalog", "Add product to cart",
        "Catalog open",
        "1. Tap product card",
        "Item added to cart, qty 1",
        "P0", "Critical", "Manual", "Passed (manual 2026-04-24)", "Confirmed in cart screenshot"),
    ("TC-053", "Catalog", "Product list caches for 5 min",
        "Products loaded once",
        "1. Reopen catalog within 5 min",
        "No network call, served from cache",
        "P2", "Performance", "Manual (check logcat)", "", ""),
    ("TC-054", "Catalog", "Clear product cache on logout",
        "Cached products",
        "1. Logout\n2. Login back\n3. Open catalog",
        "Products re-fetched from Odoo",
        "P2", "Regression", "Manual", "", ""),

    # ======================== 7. CART & ORDER ===================================
    ("TC-060", "Cart", "Increase quantity via +",
        "Item in cart",
        "1. Tap + on cart row",
        "Qty +1; subtotal updates",
        "P0", "Critical", "Manual", "", ""),
    ("TC-061", "Cart", "Decrease quantity via -",
        "Qty >= 2",
        "1. Tap - on cart row",
        "Qty -1; subtotal updates",
        "P0", "Critical", "Manual", "", ""),
    ("TC-062", "Cart", "Remove item when qty hits 0",
        "Qty = 1",
        "1. Tap - on cart row",
        "Item removed from cart",
        "P1", "Functional", "Manual", "", ""),
    ("TC-063", "Cart", "Dine In vs Application Price toggle",
        "Items in cart",
        "1. Tap 'APPLICATION PRICE'\n2. Tap 'DINE IN PRICE'",
        "Row prices switch to the corresponding pricelist; total recomputed",
        "P1", "Critical", "Manual", "", ""),
    ("TC-064", "Cart", "Subtotal = sum(price * qty)",
        "Multiple items",
        "1. Add Barotta x1 + LED x1\n2. Inspect Subtotal",
        "Subtotal = 1.000 + 110.000 = 111.000 OMR",
        "P0", "Critical", "Manual", "Passed (visual 2026-04-24)", "Cart screenshot shows OMR 111.000"),
    ("TC-065", "Cart", "Tax line computed (tax-inclusive)",
        "Items with tax config",
        "1. View Taxes row",
        "Matches Odoo pos.config tax rate * subtotal",
        "P0", "Critical", "Manual", "Passed (visual 2026-04-24)", "Taxes OMR 16.650"),
    ("TC-066", "Cart", "Total = Subtotal + Taxes",
        "Cart populated",
        "1. Subtotal 111.000 + Taxes 16.650\n2. Compare Total",
        "Total shows 127.650 OMR",
        "P0", "Critical", "Manual", "Passed (visual 2026-04-24)", "Total OMR 127.650"),
    ("TC-067", "Cart", "Order name/number from Odoo",
        "New order created",
        "1. Observe 'Restaurant/00004 Order #16'",
        "Matches pos.order.name in Odoo after create",
        "P1", "Integration", "Manual", "Passed (visual 2026-04-24)", ""),
    ("TC-068", "Cart", "Cart cleared after successful payment",
        "Pay Now success",
        "1. Complete payment\n2. Return to catalog",
        "Cart is empty",
        "P0", "Regression", "Manual", "", ""),

    # ======================== 8. PAYMENT ========================================
    ("TC-070", "Payment", "Cash payment (exact amount)",
        "Cart total 127.650",
        "1. Tap Pay Now\n2. Select Cash\n3. Enter 127.650\n4. Confirm",
        "Order created + pos.payment created; receipt printable",
        "P0", "Critical", "Manual", "", ""),
    ("TC-071", "Payment", "Cash payment with change",
        "Total 100, user enters 150",
        "1. Enter 150 in keypad\n2. Confirm",
        "Change shown = 50; Odoo stores amount_return",
        "P0", "Critical", "Manual", "", ""),
    ("TC-072", "Payment", "Card payment via bank journal",
        "Bank journal configured",
        "1. Select Card mode\n2. Enter amount\n3. Confirm",
        "pos.payment created with bank journal + paymentMethodId",
        "P0", "Critical", "Manual", "", ""),
    ("TC-073", "Payment", "Customer Account payment",
        "Selected customer has receivable",
        "1. Select Account mode\n2. Enter amount",
        "account.payment created against partner receivable",
        "P1", "Critical", "Manual", "", ""),
    ("TC-074", "Payment", "Split payment (cash + card)",
        "Half cash, half card",
        "1. Pay 60 cash\n2. Pay 67.650 card",
        "Two pos.payment records; order state = paid",
        "P1", "Critical", "Manual", "", ""),
    ("TC-075", "Payment", "Payment PIN gate blocks unauthorized user",
        "pos.config.payment_pin set",
        "1. Tap Pay Now\n2. Enter wrong PIN",
        "Payment blocked; error toast",
        "P0", "Critical", "Manual", "", ""),
    ("TC-076", "Payment", "Payment PIN gate allows with correct PIN",
        "pos.config.payment_pin set",
        "1. Tap Pay Now\n2. Enter correct PIN",
        "Proceeds to payment flow",
        "P0", "Critical", "Manual", "", ""),

    # ======================== 9. IDEMPOTENCY / DUPLICATE ORDERS ================
    ("TC-080", "Idempotency", "First create sends client_uuid",
        "Network OK; APK fabd2451+; module v19.0.2.0.0",
        "1. Pay Now normally\n2. Verify pos.order in Odoo has client_uuid set",
        "pos.order.client_uuid contains a v4 UUID (xxxxxxxx-xxxx-4xxx-yxxx-xxxxxxxxxxxx)",
        "P0", "Critical", "Manual (Odoo backend)", "Ready to test", "Field active per check_idempotency_module.py"),
    ("TC-081", "Idempotency", "Retry returns SAME order (dedup)",
        "Module v19.0.2.0.0 installed",
        "1. Pay Now\n2. Break tunnel mid-flight\n3. Reconnect\n4. Tap Retry",
        "Only ONE pos.order; Odoo log: 'duplicate create for client_uuid=... -> returning existing order id=...'",
        "P0", "Critical", "Manual", "Ready to test", "Key fix - run after fresh APK"),
    ("TC-082", "Idempotency", "New cart gets fresh client_uuid",
        "Prior order succeeded",
        "1. Complete order A\n2. Start new cart\n3. Pay order B",
        "Two distinct pos.order records with DIFFERENT client_uuid values",
        "P1", "Critical", "Manual", "Ready to test", ""),
    ("TC-083", "Idempotency", "Cancel does not consume the uuid",
        "Mid-payment",
        "1. Pay Now\n2. Network error -> Cancel\n3. Pay Now again",
        "Same uuid reused (no duplicate even if 1st reached server)",
        "P1", "Functional", "Manual", "Ready to test", ""),
    ("TC-084", "Idempotency", "client_uuid field + index on pos_order",
        "Module v19.0.2.0.0 installed",
        "1. Run check_idempotency_module.py\n2. Confirm field 'pos.order.client_uuid' = char",
        "Field type=char, label='Client UUID'; module v19.0.2.0.0 active",
        "P1", "Integration", "Python (RPC)", "Passed (script 2026-04-25)", "tools/check_idempotency_module.py"),

    # ======================== 10. NETWORK / ERROR HANDLING ======================
    ("TC-090", "Network", "Network error UI on Odoo unreachable",
        "Server truly unreachable",
        "1. Break adb reverse tunnel\n2. Tap Take Orders (or any server-bound action)",
        "Either modal 'Cannot reach server' (Retry/Cancel) or inline 'Failed to load POS registers' (Retry) appears",
        "P0", "Critical", "Appium: network-error.e2e.js", "Passed (Appium 2026-04-25)", "Inline path verified end-to-end"),
    ("TC-091", "Network", "Retry replays the same request and recovers",
        "Error UI shown, then connection restored",
        "1. Restore network\n2. Tap Retry",
        "Error UI dismisses, content loads (POS sessions visible)",
        "P0", "Critical", "Appium: network-error.e2e.js", "Passed (Appium 2026-04-25)", ""),
    ("TC-092", "Network", "Retry while still offline keeps the error visible",
        "Error UI shown, network still broken",
        "1. Tap Retry",
        "Error UI re-appears (does NOT prematurely clear)",
        "P1", "Regression", "Appium: network-error.e2e.js", "Passed (Appium 2026-04-25)", ""),
    ("TC-093", "Network", "Popup does NOT fire on 4xx/5xx with response",
        "Odoo returns 400 (validation error)",
        "1. Send intentionally bad payload",
        "Error goes to catch block directly; no popup (only thrown via error.response path)",
        "P1", "Regression", "Manual", "", ""),
    ("TC-094", "Network", "Cannot reach server on Register page even online",
        "Device config done (device_server_url set via Device Setup)",
        "1. Open Register page on good network",
        "No popup; Register loads active sessions",
        "P0", "Regression", "Manual", "Skipped (not a bug)", "Initial suspicion was ODOO_BASE_URL='' but device URL is supplied via device_server_url / AsyncStorage at runtime; behavior is correct."),

    # ======================== 11. KOT / PRINTING ================================
    ("TC-100", "KOT", "Kitchen Bill prints KOT to printer IP",
        "kot_printer_ip configured",
        "1. Add items\n2. Tap Kitchen Bill",
        "KOT printed at configured printer; Odoo logs ESC/POS call",
        "P0", "Critical", "Manual", "", ""),
    ("TC-101", "KOT", "KOT print fallback to Odoo print agent",
        "kot_use_print_agent = true",
        "1. Tap Kitchen Bill",
        "Print agent URL called, not direct TCP",
        "P1", "Integration", "Manual", "", ""),
    ("TC-102", "KOT", "KOT shows correct order_name after server assigns",
        "New order",
        "1. Pay then print",
        "Printed order number == Odoo pos.order.name",
        "P0", "Regression", "Manual", "", "Related to idempotency"),

    # ======================== 12. MULTI-DEVICE / CONCURRENCY ====================
    ("TC-110", "Concurrency", "2 APKs create orders simultaneously",
        "Both connected to same session",
        "1. Tap Pay Now on tablet A and B within 1s",
        "Odoo assigns sequential order numbers; no collision",
        "P0", "Critical", "Manual (two devices)", "", ""),
    ("TC-111", "Concurrency", "APK offline, web POS creates order, APK reconnects",
        "5 APK + 2 web same session",
        "1. APK-1 cart ready\n2. Disconnect APK-1\n3. Web creates order\n4. Reconnect APK-1\n5. Pay Now",
        "APK-1 gets NEXT available order number; no duplicate",
        "P0", "Critical", "Manual", "", ""),
    ("TC-112", "Concurrency", "Session closed by another terminal",
        "APK-1 has open cart",
        "1. Web POS closes session\n2. APK-1 taps Pay Now",
        "Graceful error shown; user instructed to start new session",
        "P1", "Critical", "Manual", "", ""),
    ("TC-113", "Concurrency", "Stock oversell two terminals sell last item",
        "product qty=1",
        "1. Both terminals add same product and Pay Now simultaneously",
        "Both succeed (documented behavior) OR one fails (if stock reservation enabled)",
        "P2", "Known limitation", "Manual", "", ""),

    # ======================== 13. LOGOUT & SESSION PERSISTENCE =================
    ("TC-120", "Logout", "Logout clears session",
        "Logged in",
        "1. Tap Logout",
        "Returns to Login screen; userData + session_id removed",
        "P0", "Critical", "Manual", "", ""),
    ("TC-121", "Logout", "Saved credentials persist after logout",
        "Autofill was ON",
        "1. Logout\n2. Observe Login screen",
        "saved_credentials still usable if autofill toggled ON again",
        "P2", "Functional", "Manual", "", ""),

    # ======================== 14. ACCESSIBILITY / UI POLISH =====================
    ("TC-130", "UI", "Orientation - portrait vs landscape",
        "Any screen",
        "1. Rotate tablet",
        "Layout adapts; no clipped UI",
        "P2", "UI", "Manual", "", ""),
    ("TC-131", "UI", "System back button from cart returns to catalog",
        "Cart open",
        "1. Press Android back",
        "Goes to product catalog; cart state preserved",
        "P1", "Regression", "Manual", "", ""),
    ("TC-132", "UI", "Toast messages auto-dismiss",
        "Any toast",
        "1. Trigger toast\n2. Wait 3s",
        "Toast disappears",
        "P2", "UI", "Manual", "", ""),

    # ======================== 15. PERFORMANCE / NON-FUNCTIONAL =================
    ("TC-140", "Perf", "Login < 3s on good network",
        "Good WiFi",
        "1. Time from Login tap to Home render",
        "< 3 seconds",
        "P2", "Perf", "Manual (timer)", "", ""),
    ("TC-141", "Perf", "Product list loads < 2s from cache",
        "Products cached",
        "1. Reopen catalog",
        "< 2 seconds",
        "P2", "Perf", "Manual", "", ""),
    ("TC-142", "Perf", "No memory leak after 50 orders",
        "Long session",
        "1. Place 50 orders back-to-back\n2. Check Android Memory",
        "No continuous growth past steady state",
        "P2", "Perf", "Manual", "", ""),

    # ======================== 16. CURRENCY / LOCALIZATION (Oman / OMR) ============
    ("TC-150", "Currency", "Company currency configured as OMR",
        "Odoo company set up for Oman",
        "1. Odoo Settings -> Companies -> select company\n2. Inspect Currency field",
        "Currency = OMR (Omani Rial), 3-decimal precision",
        "P0", "Localization", "Manual (Odoo backend)", "Passed (script 2026-04-25)", "tools/setup_oman_pos.py applied"),
    ("TC-151", "Currency", "Product prices stored in OMR with 3 decimals",
        "Currency = OMR",
        "1. Open product master in Odoo\n2. Verify Sales Price field precision",
        "Price 1.000, 4.000, 110.000 — never 1.00 (2-decimal USD format)",
        "P0", "Localization", "Manual (Odoo backend)", "Passed (script 2026-04-25)", "OMR rounding=0.001"),
    ("TC-152", "Currency", "Tax rate configured for Oman VAT (5%)",
        "account.tax record exists",
        "1. Odoo -> Accounting -> Configuration -> Taxes\n2. Inspect default sales tax",
        "Tax = 5% VAT (price_include flag based on inclusive vs exclusive policy)",
        "P0", "Localization", "Manual (Odoo backend)", "Passed (script 2026-04-25)", "Tax id=1 -> VAT 5% (Oman)"),
    ("TC-153", "Currency", "Cart subtotal / tax / total shown in OMR (3 decimals)",
        "Items in cart",
        "1. Add items in APK\n2. Inspect Subtotal / Taxes / Total rows",
        "All amounts use OMR prefix and 3-decimal format (e.g. 'OMR 127.650')",
        "P0", "Critical", "Manual", "Passed (visual 2026-04-24)", "Cart screenshot showed OMR 5.000 / 0.750 / 5.750 correctly"),
    ("TC-154", "Currency", "Receipt prints OMR code or symbol",
        "Order completed",
        "1. Pay an order\n2. Print receipt",
        "Receipt shows OMR or Omani Rial — never $/USD",
        "P0", "Localization", "Manual (printer)", "Not run", ""),
    ("TC-155", "Currency", "Odoo pos.order amount_total stored in OMR",
        "Order completed",
        "1. Inspect pos.order in Odoo backend\n2. Look at currency_id and amount_total fields",
        "currency_id = OMR record id; amounts match cart math",
        "P1", "Integration", "Manual (Odoo backend)", "Not run", ""),
    ("TC-156", "Currency", "Local Odoo dev instance has restaurant POS products seeded",
        "Local Odoo at localhost:8069",
        "1. Open POS UI in Odoo\n2. Confirm restaurant menu items exist (Barotta, Fried Rice, etc.)",
        "Products visible with OMR prices and OMR-based taxes assigned",
        "P1", "Setup", "Manual", "Passed (script 2026-04-25)", "11 products seeded via setup_oman_pos.py"),

    # ======================== 17. SPEED / PERFORMANCE CHECKLIST (manual stopwatch) ====
    ("TC-160", "Speed", "Cold app launch -> Splash -> Login screen",
        "App freshly installed; tap icon",
        "1. Start stopwatch as you tap icon\n2. Stop when 'Welcome Back' is fully drawn",
        "< 4 seconds on Galaxy Tab A; < 3s on flagship phone",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-161", "Speed", "Warm app launch (already logged in) -> Home",
        "User logged in; kill app from recents; reopen",
        "1. Stopwatch from icon tap\n2. Stop when Home dashboard 'Take Orders' visible",
        "< 3 seconds",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-162", "Speed", "Login submit -> Home dashboard",
        "On Login screen with credentials filled",
        "1. Tap Login\n2. Stop when 'Take Orders' banner visible",
        "< 3 seconds on good WiFi",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-163", "Speed", "Take Orders -> POS Register list",
        "Logged in",
        "1. Tap 'Take Orders'\n2. Stop when Active Sessions list fully rendered",
        "< 2 seconds",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-164", "Speed", "Continue Selling -> Choose Order Type screen",
        "POS Register screen",
        "1. Tap 'Continue Selling'\n2. Stop when 'How would you like to serve today?' visible",
        "< 2 seconds",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-165", "Speed", "Dine In -> Select Table screen (Main Floor + 12 tables drawn)",
        "Choose Order Type screen",
        "1. Tap Dine In\n2. Stop when all table cards (T1-T12) rendered",
        "< 3 seconds",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-166", "Speed", "Tap table -> POS Products with category list",
        "Select Table screen",
        "1. Tap T2\n2. Stop when product grid + DINE IN PRICE visible",
        "< 4 seconds (cold) / < 1s (cached)",
        "P0", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-167", "Speed", "Tap Food category -> products filtered",
        "POS Products screen",
        "1. Tap a category chip\n2. Stop when filtered grid visible",
        "< 1 second",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-168", "Speed", "Tap product -> appears in cart",
        "POS Products screen",
        "1. Tap Karak Tea\n2. Stop when item shown in cart panel",
        "< 500ms",
        "P0", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-169", "Speed", "Pay Now -> payment method modal",
        "Cart with at least 1 item",
        "1. Tap Pay Now\n2. Stop when 'Select Payment Method' modal renders",
        "< 1.5 seconds",
        "P0", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-170", "Speed", "Confirm Cash payment -> success toast / receipt",
        "Payment modal, cash selected, amount entered",
        "1. Tap confirm\n2. Stop when Order Created/Paid feedback visible",
        "< 5 seconds (writes pos.order + pos.payment)",
        "P0", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-171", "Speed", "Kitchen Bill button -> KOT preview screen",
        "Cart populated",
        "1. Tap Kitchen Bill\n2. Stop when item list visible",
        "< 1 second",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-172", "Speed", "Print Full Order -> KOT printed",
        "KOT preview, printer reachable",
        "1. Tap Print Full Order\n2. Stop when printer ejects ticket",
        "< 6 seconds (network + printer)",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-173", "Speed", "New Takeout Order flow -> order created",
        "Choose Order Type",
        "1. Tap New Takeout Order\n2. Enter customer name + slot\n3. Stop when products screen ready",
        "< 4 seconds",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-174", "Speed", "Logout -> Login screen",
        "Logged in, on Home",
        "1. Tap Logout in bottom nav\n2. Confirm dialog\n3. Stop when 'Welcome Back' visible",
        "< 1.5 seconds",
        "P2", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-175", "Speed", "Network popup recovery (Retry after WiFi back on)",
        "Cannot reach server popup visible",
        "1. Re-enable WiFi\n2. Tap Retry\n3. Stop when target screen loads",
        "< 4 seconds after network restored",
        "P0", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-176", "Speed", "Search product by name",
        "POS Products screen",
        "1. Type 'Bar' in search\n2. Stop when only Barotta visible",
        "< 500ms (in-memory filter)",
        "P1", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-177", "Speed", "Bottom nav switch (Home -> Profile -> Logout)",
        "Logged in",
        "1. Tap each nav tab in succession",
        "Each transition < 500ms; no flicker",
        "P2", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-178", "Speed", "Language switch EN <-> AR re-render",
        "Login or any screen",
        "1. Tap عربي (Arabic)\n2. Stop when all text translated + RTL flipped",
        "< 1 second",
        "P2", "Perf", "Manual stopwatch", "Not run", ""),
    ("TC-179", "Speed", "Idle memory growth check (50 product taps)",
        "POS Products screen",
        "1. Tap product 50 times in a row\n2. Watch tablet RAM in Settings -> Battery",
        "Memory should plateau, not grow linearly",
        "P2", "Perf", "Manual (Settings)", "Not run", "Memory leak detection"),
]


def build_wb():
    wb = Workbook()

    # === Summary sheet ======================================================
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "NexGenn Restaurant POS - Test Scenarios"
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="2E294E")
    summary.merge_cells("A1:C1")
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 35

    # counts
    total = len(SCENARIOS)
    by_pri = {"P0": 0, "P1": 0, "P2": 0}
    by_mod = {}
    for r in SCENARIOS:
        by_pri[r[6]] = by_pri.get(r[6], 0) + 1
        by_mod[r[1]] = by_mod.get(r[1], 0) + 1

    summary["A3"] = "Total scenarios"
    summary["B3"] = total
    summary["A4"] = "P0 (critical)"
    summary["B4"] = by_pri["P0"]
    summary["A5"] = "P1 (important)"
    summary["B5"] = by_pri["P1"]
    summary["A6"] = "P2 (nice-to-have)"
    summary["B6"] = by_pri["P2"]

    for r in range(3, 7):
        summary.cell(row=r, column=1).font = Font(bold=True)

    # Status rollup
    def bucket(s):
        s = (s or "").lower()
        if not s or "not run" in s:
            return "Not run"
        if "skip" in s:
            return "Skipped"
        if "pass" in s:
            return "Passed"
        if "bug" in s or "fail" in s:
            return "Failed / Open bug"
        if "block" in s:
            return "Blocked"
        if "ready" in s:
            return "Ready to test"
        return "Other"
    by_status = {}
    for r in SCENARIOS:
        status = r[9] if len(r) >= 10 else ""
        b = bucket(status)
        by_status[b] = by_status.get(b, 0) + 1

    summary["A8"] = "By Status"
    summary["A8"].font = Font(bold=True, size=12)
    status_colours = {
        "Passed":            "C8E6C9",
        "Failed / Open bug": "FFCDD2",
        "Blocked":           "FFE0B2",
        "Ready to test":     "BBDEFB",
        "Not run":           "ECEFF1",
        "Skipped":           "E1BEE7",
        "Other":             "FFFFFF",
    }
    status_order = ["Passed", "Failed / Open bug", "Blocked", "Ready to test", "Not run", "Skipped", "Other"]
    row_ix = 9
    for st in status_order:
        if by_status.get(st):
            c = summary.cell(row=row_ix, column=1, value=st)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=status_colours[st])
            summary.cell(row=row_ix, column=2, value=by_status[st])
            row_ix += 1

    # By module after a spacer
    row_ix += 1
    summary.cell(row=row_ix, column=1, value="By Module").font = Font(bold=True, size=12)
    row_ix += 1
    for (mod, cnt) in sorted(by_mod.items()):
        summary.cell(row=row_ix, column=1, value=mod).font = Font(bold=True)
        summary.cell(row=row_ix, column=2, value=cnt)
        row_ix += 1

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 10

    # === Numbering scheme sheet =============================================
    ns = wb.create_sheet("Numbering Scheme")
    ns["A1"] = "TC ID Numbering Scheme"
    ns["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ns["A1"].fill = PatternFill("solid", fgColor="2E294E")
    ns["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ns.merge_cells("A1:C1")
    ns.row_dimensions[1].height = 32

    ns["A3"] = (
        "IDs are allocated in per-module 10-blocks on purpose. Gaps like TC-006/007/008/009 "
        "are RESERVED SLOTS for future tests in that module. Adding a new Login test later "
        "becomes TC-017 without shifting any existing ID. This is standard QA practice."
    )
    ns["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ns.merge_cells("A3:C4")
    ns.row_dimensions[3].height = 30
    ns.row_dimensions[4].height = 30

    # Header
    hdr = [("ID Range", 14), ("Module", 28), ("Notes", 55)]
    for col, (name, width) in enumerate(hdr, start=1):
        c = ns.cell(row=6, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E294E")
        c.alignment = Alignment(horizontal="center")
        ns.column_dimensions[get_column_letter(col)].width = width
    ns.row_dimensions[6].height = 22

    blocks = [
        ("TC-001 to TC-009", "App Launch",      "Splash, warm/cold start, icon launch"),
        ("TC-010 to TC-019", "Login",           "Credentials, validation, password toggle, autofill"),
        ("TC-020 to TC-029", "i18n (EN / AR)",  "Language switch, RTL flip, persistence"),
        ("TC-030 to TC-039", "Home",            "Dashboard, Take Orders banner, Our Specials"),
        ("TC-040 to TC-049", "POS Register",    "Active sessions, Continue Selling, Close"),
        ("TC-050 to TC-059", "Catalog",         "Products, search, add to cart, cache"),
        ("TC-060 to TC-069", "Cart & Order",    "Qty, price toggle, subtotal, tax, total"),
        ("TC-070 to TC-079", "Payment",         "Cash, card, account, split, PIN gate"),
        ("TC-080 to TC-089", "Idempotency",     "pos_reference UUID, duplicate prevention"),
        ("TC-090 to TC-099", "Network / Error", "Popup, retry, cancel, error classification"),
        ("TC-100 to TC-109", "KOT / Printing",  "Kitchen bill, print agent, order name"),
        ("TC-110 to TC-119", "Concurrency",     "5 APK + 2 web conflicts, session locks"),
        ("TC-120 to TC-129", "Logout / Session","Logout, saved credentials persistence"),
        ("TC-130 to TC-139", "UI / UX",         "Orientation, back button, toasts"),
        ("TC-140 to TC-149", "Performance",     "Login time, cache speed, memory"),
        ("TC-150 to TC-159", "Currency / L10n", "OMR currency, 3 decimals, 5% VAT, OMR receipts (Oman)"),
        ("TC-160 to TC-179", "Speed Checklist", "Manual stopwatch tests for every key tap-to-render"),
    ]

    thin = Side(border_style="thin", color="CCCCCC")
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (rng, module, note) in enumerate(blocks, start=7):
        ns.cell(row=i, column=1, value=rng).font = Font(bold=True, name="Consolas")
        ns.cell(row=i, column=2, value=module)
        ns.cell(row=i, column=3, value=note)
        for col in range(1, 4):
            c = ns.cell(row=i, column=col)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = border_cell
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor="F5F5F5")

    # Rules of thumb
    start_rules = len(blocks) + 9
    ns.cell(row=start_rules, column=1, value="Rules of thumb").font = Font(bold=True, size=12)
    rules = [
        "- A block has 10 slots. 5-6 used now, 4-5 reserved for future.",
        "- If a block fills up, extend with 2xx / 3xx continuation (e.g. TC-210+).",
        "- Never reuse an ID for a different scenario. Retire old IDs as 'Removed'.",
        "- When adding a new scenario, pick the lowest free slot inside the matching block.",
    ]
    for i, rule in enumerate(rules, start=start_rules + 1):
        ns.cell(row=i, column=1, value=rule)
        ns.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        ns.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # === Test cases sheet ===================================================
    ws = wb.create_sheet("Test Cases")

    # Header
    for col, (name, width) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="2E294E")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30

    # Priority colour map
    pri_fill = {
        "P0": PatternFill("solid", fgColor="FFCDD2"),   # red
        "P1": PatternFill("solid", fgColor="FFE0B2"),   # orange
        "P2": PatternFill("solid", fgColor="DCEDC8"),   # green
    }
    # Status colour map
    def status_fill(status):
        s = (status or "").lower()
        if "skip" in s:
            return PatternFill("solid", fgColor="E1BEE7")   # purple
        if "pass" in s:
            return PatternFill("solid", fgColor="C8E6C9")   # green
        if "bug" in s or "fail" in s:
            return PatternFill("solid", fgColor="FFCDD2")   # red
        if "block" in s:
            return PatternFill("solid", fgColor="FFE0B2")   # orange
        if "ready" in s:
            return PatternFill("solid", fgColor="BBDEFB")   # blue
        if "not run" in s:
            return PatternFill("solid", fgColor="ECEFF1")   # grey
        return None

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row in enumerate(SCENARIOS, start=2):
        row = list(row)
        # Default empty Status (column 10) to "Not run"
        if len(row) >= 10 and (row[9] == "" or row[9] is None):
            row[9] = "Not run"
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if col == 7 and val in pri_fill:  # Priority column
                cell.fill = pri_fill[val]
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col == 10:  # Status column
                fill = status_fill(val)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)

        # Zebra striping (skip Priority and Status columns which have their own fills)
        if r % 2 == 0:
            for col in range(1, len(COLS) + 1):
                if col not in (7, 10):
                    cur = ws.cell(row=r, column=col)
                    if cur.fill.start_color.rgb in (None, "00000000"):
                        cur.fill = PatternFill("solid", fgColor="F5F5F5")

    # Freeze header row + ID column
    ws.freeze_panes = "B2"

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(SCENARIOS) + 1}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}  ({total} scenarios)")


if __name__ == "__main__":
    build_wb()
